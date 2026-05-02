#Scrapes the 2023-2024 Premier League standing from Wikipedia. 

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl import Font, PatternFill, Alignment, Border, Side
import os

URL = 'https://en.wikipedia.org/wiki/2023%E2%80%9324_Premier_League'

HEADERS = {'User-Agent' : ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')}
SOURCE = ('Wikipedia (2024) 2023-2024 Premier League.''Available at: https://en.wikipedia.org/wiki/2023%E2%80%9324_Premier_League')
def scrape_standings():
    print("Scraping Premier League standings from Wikipedia...")
    response = requests.get(URL, headers=HEADERS, timeout=15)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, 'html.parser')
    #Find the League table
    table = None 
    for t in soup.find_all('table', {'class': 'wikitable'}): 
        header = t.find('tr')
        if header and 'Pos' in header.text and 'GF' in header.text and 'Pts' in header.text:
            table = t
            break

    if not table:
        raise ValueError("Could not find the league standings table on the Wikipedia page.")    
    rows = table.find('tbody').find_all('tr')
    Clubs = []

    import re

    #Identify the correct club names
    name_lookup_table = {
        'Manchester City': 'Manchester City',
        'Arsenal': 'Arsenal',
        'Newcastle United': 'Newcastle United',
        'Liverpool': 'Liverpool',
        'Brighton & Hove Albion': 'Brighton & Hove Albion',
        'Fulham': 'Fulham',
        'Aston Villa': 'Aston Villa',
        'Tottenham Hotspur': 'Tottenham Hotspur',
        'Brentford': 'Brentford',
        'Crystal Palace': 'Crystal Palace',
        'Chelsea': 'Chelsea',
        'West Ham United': 'West Ham United',
        'Bournemouth': 'AFC Bournemouth',
        'Nottingham Forest': 'Nottingham Forest',
        'Everton': 'Everton',
        'Leicester City': 'Leicester City',
        'Wolverhampton Wanderers': 'Wolverhampton Wanderers',
        "Luton Town": "Luton Town",
        "Sheffield United": "Sheffield United",

    }
    for row in rows:
        cells = row.find_all(['td', 'th'])
        if len(cells) > 9:
            continue

        try:
            position = int(cells[0].text.strip())
            #Remove suffixes and footnotes from the club names
            raw_club_name = cells[1].text.strip()
            raw_club_name = re.sub(r'\[.*?\]', '', raw_club_name).strip()
            Club = name_lookup_table.get(raw_club_name, raw_club_name)
            gf = int(cells[6].text.strip())
            ga = int(cells[7].text.strip())
            #Remove suffixes and footnotes from the points
            raw_points = cells[9].text.strip()
            pts = int(re.sub(r'\[.*?]', '', raw_points))
            Clubs.append((position, Club, gf, ga))
        except (ValueError, IndexError):
            continue

    if len(Clubs) != 20:
        raise ValueError(f"Expected to scrape 20 clubs, but got {len(Clubs)}. Check the Wikipedia page structure.")
    print(f"Successfully scraped standings for {len(Clubs)} clubs.")
    return Clubs

def save_to_excel(clubs):
    xl_path = 'data/PL_Finances_VS_Performance.xlsx'
    wb = openpyxl.load_workbook(xl_path)

    if 'scraped_standings' in wb.sheetnames:
        del wb['scraped_standings']
    ws = wb.create_sheet('scraped_standings')

    header_font = Font(bold=True, colour= 'FFFFFF', size=11)
    header_fill= PatternFill('solid', fgColor='#3776ab')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center = Alignment(horizontal='center', vertical='center')
    headers = ['League Position', 'Club', 'Final Points', 'Goals Scored', 'Goals Conceded', 'Source']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    for row, (pos, Club, pts, gf, ga) in enumerate(clubs, 2):
        for col, val in enumerate([pos, Club, pts, gf, ga, SOURCE], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            cell.alignment = center
   
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 80

    wb.save(xl_path)
    print(f"Scraped standings saved to '{xl_path}' - sheet: scraped_standings")

if __name__ == "__main__":
    os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    clubs = scrape_standings()
    save_to_excel(clubs)

    print("\nScraped stndings preview:")
    print(f"{'Pos':<5} {'Club':<30} {'Pts':<6} {'GF':<6} {'GA':<6}")
    print("-" * 55)
    for pos, Club, pts, gf, ga in clubs:
        print(f"{pos:<5} {Club:<30} {pts:<6} {gf:<6} {ga:<6}")